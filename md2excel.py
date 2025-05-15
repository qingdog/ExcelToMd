import re

import openpyxl
from openpyxl.styles import Alignment, Font  # 导入字体和对齐方式


def md2excel():
    # 创建一个Excel工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 读取markdown文件内容
    file_path = "1xmind导出成markdown.md"  # markdown文件路径
    with open(file_path, encoding='utf-8') as file:
        lines = file.readlines()

    # 处理读取到的markdown内容，并将其写入Excel
    row = 0  # excel行数
    previous_first = ""  # 记录需求一行
    i = 0  # 文件行索引

    excel_title = "相关研发需求	用例标题	前置条件	步骤	预期	实际情况	优先级	用例类型"
    # titles = excel_title.split(sep=None, maxsplit=-1)
    titles = re.split(r'[\s,，]', excel_title)
    row += 1
    column = 0
    for title in titles:
        column = insert_column(row, column, title, ws)

    cycle_index = 0
    while i < len(lines) - 1:
        cycle_index += 1

        line, i = continue_line(lines, i)
        if not line: break

        # 新的一行
        row += 1
        column = 0

        # 跳过标题
        if re.search("^# ", line) is not None:
            i += 1
            pass

        # 相关研发需求
        line, i = continue_line(lines, i)
        if not line: break
        if re.search(r"^## ", line) is not None:
            i += 1
            line = re.sub(r"^## ", "", line, count=1)
            column = insert_column(row, column, line, ws)
            previous_first = line
            column = 1
        else:
            column = insert_column(row, column, previous_first, ws)
            column = 1
            # i -= 1

        # 用例标题
        line, i = continue_line(lines, i)
        print(f"=================={line} {i}")
        if not line: break
        if re.search(r"^### ", line) is not None:
            i += 1
            line = re.sub(r"^### ", "", line, count=1)
            column = insert_column(row, column, line, ws)

        # 前置条件
        line, i = continue_line(lines, i)
        if not line: break
        if re.search(r"^#### ", line) is not None:
            i += 1
            line = re.sub(r"^#### ", "", line, count=1)
            column = insert_column(row, column, line, ws)

        # 步骤
        line, i = continue_line(lines, i)
        if not line: break
        if re.search("^- ", line) is not None:
            i += 1
            steps = line.replace("- ", "")

            # print(123)
            while True:
                line1, ii = continue_line(lines, i)
                if not line1: break
                if re.search("^- ", line1) is not None:
                    i += 1
                    steps += line1.replace("- ", "\n")
                    # i = ii
                else:
                    break
            # print(456)
            # print(f"======================================={i}")
            # print(f"======================================={i}")
            column = insert_column(row, column, steps, ws)

        # 预期
        # print(f"======================================={i}")
        line, i = continue_line(lines, i)
        if not line: break
        if re.search("^ {2}- ", line) is not None:
            i += 1
            expects = line.replace("  - ", "")
            print("expects===" + expects)

            while True:
                line1, ii = continue_line(lines, i)
                if not line1: break
                if re.search("^ {2}- ", line1) is not None:
                    i += 1
                    expects += line1.replace("  - ", "\n")
                    # i = ii
                else:
                    break

            column = insert_column(row, column, expects, ws)

    # 设置固定列宽
    column_widths = [30, 30, 16, 80, 70]  # 你指定的列宽
    for i, width in enumerate(column_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)  # 获取Excel列的字母
        ws.column_dimensions[col_letter].width = width

    # 设置行高（可选）
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 70  # 统一设置20像素高度

    # 保存为Excel文件
    output_file = "1markdown转成的excel.xlsx"
    wb.save(output_file)
    print(f"Excel文件 '{output_file}' 已成功创建！")


def continue_line(lines, i):
    while i < len(lines) - 1:
        # 处理每一行
        line = lines[i]
        line = re.sub(r"\n$", "", line)
        if not line or re.search(r"^[\s#]$", line) is not None:
            i += 1
            continue
        print(f"{line} {i}")
        # i += 1 # 获取完下一行后，直接加1
        return line, i
    return None, i


def insert_column(row, column, line, ws):
    column += 1
    cell = ws.cell(row=row, column=column, value=line)  # 将内容写入Excel的第一列
    # 设置字体、大小、自动换行和垂直居中
    cell.font = Font(name="微软雅黑", size=12)  # 设置字体为微软雅黑，大小12
    cell.alignment = Alignment(wrapText=True, vertical="center")  # 自动换行+垂直居中

    return column


if __name__ == '__main__':
    md2excel()
